
# First part is to create required Agro files

import os
import pandas as pd
from datetime import datetime
import yaml

# Custom representer to format dates correctly without time
def represent_date(self, data):
    value = data.strftime('%Y-%m-%d')
    return self.represent_scalar('tag:yaml.org,2002:str', value)

yaml.add_representer(datetime.date, represent_date)

def create_agro_yaml_file(index, id, sowing_date, harvest_date, directory):
    campaign_start_date = datetime(sowing_date.year, 1, 1).date()  # Ensure it's a date only
    data = {
        'AgroManagement': [{
            campaign_start_date: {
                'CropCalendar': {
                    'crop_name': 'relevant-crop',
                    'variety_name': 'relevant_crop_code',
                    'crop_start_date': sowing_date,
                    'crop_start_type': 'sowing',
                    'crop_end_date': harvest_date,
                    'crop_end_type': 'harvest',
                    'max_duration': 300
                },
                'TimedEvents': None,
                'StateEvents': None
            }
        }]
    }

    # Include index in the filename to ensure uniqueness
    filename = os.path.join(directory, f'agrorelevantcrop_ID_{id}_row_{index + 1}.agro')
    with open(filename, 'w') as file:
        yaml.dump(data, file, default_flow_style=False, allow_unicode=True, sort_keys=False, indent=2)
    print(f"File saved: {filename}")

# Setup paths and load data
crop_info_path = r'C:\data copy\WOFOST MODEL\Model input files\crop and agro\Crop info ID.xlsx'
output_directory = r'C:\data copy\WOFOST MODEL\Model input files\crop and agro\yaml agro files'
if not os.path.exists(output_directory):
    os.makedirs(output_directory)

crop_info = pd.read_excel(crop_info_path)
crop_info['Sowing date'] = pd.to_datetime(crop_info['Sowing date'], format='%d.%m.%Y').dt.date
crop_info['Harvest date'] = pd.to_datetime(crop_info['Harvest date'], format='%d.%m.%Y').dt.date

# Generate YAML files for each row entry
for index, row in crop_info.iterrows():
    create_agro_yaml_file(index, row['ID'], row['Sowing date'], row['Harvest date'], output_directory)

# Second part is to run the model

import os
import pandas as pd
import numpy as np
from pcse.fileinput import CABOFileReader, YAMLAgroManagementReader
from pcse.util import WOFOST71SiteDataProvider
from pcse.base import ParameterProvider
from pcse.db import NASAPowerWeatherDataProvider
from pcse.models import Wofost71_WLP_FD
import matplotlib.pyplot as plt

# Define the directories
soil_location_dir = r'C:\data copy\WOFOST MODEL\Model input files\Soil and Location'
agro_files_dir = r'C:\data copy\WOFOST MODEL\Model input files\crop and agro\yaml agro files'
id_soil_location_path = os.path.join(soil_location_dir, 'ID SOIL and Location.xlsx')
crop_info_path = os.path.join(soil_location_dir, 'Crop info ID.xlsx')
crop_parameters_file = os.path.join(agro_files_dir, '..', 'practice_relevantcrop.crop')

# Load the soil and crop information
id_soil_location = pd.read_excel(id_soil_location_path)
crop_info = pd.read_excel(crop_info_path)
crop_parameters = CABOFileReader(crop_parameters_file)

# Prepare additional columns for RRMSE
crop_info['RRMSE TWSO %'] = np.nan
crop_info['RRMSE TAGP %'] = np.nan

# Define functions
def calculate_rrmse(observed, estimated):
    observed = np.array(observed, dtype=float)
    estimated = np.array(estimated, dtype=float)
    if observed.size == 0 or estimated.size == 0 or np.isnan(observed).all() or np.isnan(estimated).all():
        return np.nan
    return np.sqrt(np.nanmean((observed - estimated) ** 2)) / np.nanmean(observed) * 100

# Errors arrays for later calculation of mean error
errors_twso = []
errors_tagp = []

# Main processing loop
for index, crop_row in crop_info.iterrows():
    id_number = crop_row['ID']
    row_number = index + 1
    agro_file_path = os.path.join(agro_files_dir, f'agrorelevantcrop_ID_{id_number}_row_{row_number}.agro')

    if not os.path.exists(agro_file_path):
        print(f"Agro file {agro_file_path} does not exist for ID {id_number}, row {row_number}.")
        continue

    
    soil_info = id_soil_location[id_soil_location['ID'] == id_number].iloc[0]

    soil_file = os.path.join(soil_location_dir, f"{soil_info['Soil File']}")
    if not os.path.exists(soil_file):
        print(f"Soil file {soil_file} does not exist for ID {id_number}, row {row_number}.")
        continue

    # Print out soil file, agro file, latitude, and longitude for each specific row
    print(f"For ID {id_number}, row {row_number}:")
    print(f"- Soil file: {soil_file}")
    print(f"- Agro file: {agro_file_path}")
    print(f"- Latitude: {soil_info['Latitude']}")
    print(f"- Longitude: {soil_info['Longitude']}")

    soildata = CABOFileReader(soil_file)
    sitedata = WOFOST71SiteDataProvider(WAV=100)
    parameters = ParameterProvider(soildata=soildata, cropdata=crop_parameters, sitedata=sitedata)
    wdp = NASAPowerWeatherDataProvider(latitude=soil_info['Latitude'], longitude=soil_info['Longitude'])
    agromanagement = YAMLAgroManagementReader(agro_file_path)
    wofsim = Wofost71_WLP_FD(parameters, wdp, agromanagement)

    try:
        wofsim.run_till_terminate()
        df = pd.DataFrame(wofsim.get_output())
        # Print all output variables for each model run
        print(f"Output variables for ID {id_number}, row {row_number}:")
        print(df.to_string(index=False))
    except Exception as e:
        print(f"Error running model for ID {id_number} row {row_number}: {e}")
        continue

    estimated_yield = df['TWSO'].iloc[-1]
    estimated_biomass = df['TAGP'].iloc[-1]
    crop_info.at[index, 'Dry matter yield (TWSO) kg/ha (estimated)'] = estimated_yield
    crop_info.at[index, 'Total Above-ground Dry Matter (TAGP) Biomass Estimated (Kg/ha)'] = estimated_biomass

    # Calculate RRMSE and store errors
    observed_yield = crop_row['Dry matter yield (TWSO) kg/ha (observed)']
    observed_biomass = crop_row['Total Above-ground Dry Matter (TAGP) Biomass observed (Kg/ha)']
    rrms_twso = calculate_rrmse([observed_yield], [estimated_yield])
    rrms_tagp = calculate_rrmse([observed_biomass], [estimated_biomass])
    crop_info.at[index, 'RRMSE TWSO %'] = rrms_twso
    crop_info.at[index, 'RRMSE TAGP %'] = rrms_tagp

    errors_twso.append(np.abs(observed_yield - estimated_yield))
    errors_tagp.append(np.abs(observed_biomass - estimated_biomass))

    # Save figures
    fig, axes = plt.subplots(nrows=2, ncols=2, figsize=(10, 8))
    for var, ax in zip(["DVS", "TAGP", "LAI", "TWSO"], axes.flatten()):
        ax.plot(df['day'], df[var], 'b-')
        ax.set_title(var)
    fig.autofmt_xdate()
    fig_path = os.path.join(agro_files_dir, f'relevantcrop_{id_number}_row_{row_number}.png')
    fig.savefig(fig_path)
    plt.close(fig)

# Calculate the mean of all RRMSE values for TWSO and TAGP
mean_rrmse_twso = crop_info['RRMSE TWSO %'].mean()
mean_rrmse_tagp = crop_info['RRMSE TAGP %'].mean()

# Create a DataFrame to hold these mean values
mean_errors_df = pd.DataFrame({
    'RRMSE TWSO %': ['Mean RRMSE TWSO %'],
    'RRMSE TAGP %': ['Mean RRMSE TAGP %'],
    'Dry matter yield (TWSO) kg/ha (estimated)': [mean_rrmse_twso],
    'Total Above-ground Dry Matter (TAGP) Biomass Estimated (Kg/ha)': [mean_rrmse_tagp]
})

# Append this new DataFrame to the existing crop_info DataFrame
crop_info = pd.concat([crop_info, mean_errors_df], ignore_index=True)

# Optionally, save the updated crop data to an Excel file
updated_crop_info_path = os.path.join(agro_files_dir, '..', 'Output_data.xlsx')
try:
    crop_info.to_excel(updated_crop_info_path, index=False)
    print("Data saved successfully. Mean errors appended.")
except PermissionError as e:
    print(f"Failed to save updated crop data due to a permission error: {e}")
except Exception as e:
    print(f"An error occurred when saving the data: {e}")

# Third part is to get the outliers

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import numpy as np

# Load the Excel file
excel_file_path = r"C:\data copy\WOFOST MODEL\Model input files\crop and agro\Output_data.xlsx"
df = pd.read_excel(excel_file_path)

# Exclude the last row containing non-numeric values
df = df.iloc[:-1]

# Define a function to find outliers
def find_outlier(data, threshold=30):
    outliers = []
    for value in data:
        if pd.notna(value) and float(value) > threshold:  # Convert to float before comparison
            outliers.append(True)
        else:
            outliers.append(False)
    return outliers

# Find outliers in columns 'RRMSE TWSO %' and 'RRMSE TAGP %'
outliers_TWSO = find_outlier(df['RRMSE TWSO %'])
outliers_TAGP = find_outlier(df['RRMSE TAGP %'])

# Calculate mean excluding outliers
mean_TWSO = np.mean([df['RRMSE TWSO %'].iloc[i] for i in range(len(outliers_TWSO)) if not outliers_TWSO[i]])
mean_TAGP = np.mean([df['RRMSE TAGP %'].iloc[i] for i in range(len(outliers_TAGP)) if not outliers_TAGP[i]])

# Load the Excel workbook using openpyxl
wb = load_workbook(excel_file_path)
ws = wb.active

# Highlight outliers in column 'RRMSE TWSO %'
for idx, outlier in enumerate(outliers_TWSO, start=2):  # Start from row 2 (assuming row 1 is header)
    if outlier:
        ws[f'H{idx}'].fill = PatternFill(start_color="FF6347", fill_type="solid")  # Highlight in red

# Highlight outliers in column 'RRMSE TAGP %'
for idx, outlier in enumerate(outliers_TAGP, start=2):  # Start from row 2 (assuming row 1 is header)
    if outlier:
        ws[f'I{idx}'].fill = PatternFill(start_color="FF6347", fill_type="solid")  # Highlight in red

# Add a new row with calculated means at the end of the existing data
last_row = ws.max_row + 1
ws[f'H{last_row}'] = mean_TWSO
ws[f'I{last_row}'] = mean_TAGP

# Save the modified Excel file
wb.save("Output_data_with_outliers.xlsx")

